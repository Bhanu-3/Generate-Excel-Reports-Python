from openpyxl import load_workbook
from openpyxl import chart
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import os
import sys

app_path = os.getcwd()
input_path = os.path.join(app_path,'pivot_table.xlsx')

month= input('Enter Month: ')

wb = load_workbook(input_path)
sheet = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

Barchart = chart.BarChart()

data = chart.Reference(sheet,
         min_col=min_col+1,
         max_col=max_col,
         min_row=min_row,
         max_row=max_row)
category = chart.Reference(sheet,
         min_col=min_col,
         max_col=min_col,
         min_row=min_row,
         max_row=max_row)

Barchart.add_data(data,titles_from_data=True)
Barchart.set_categories(category)
sheet.add_chart(Barchart,"B12")
Barchart.title='SalesByPrdLine'
Barchart.style=5

for i in range(min_col+1,max_col+1):
    let = get_column_letter(i)
    sheet[f'{let}{max_row+1}'] = f'=SUM({let}{min_row+1}:{let}{max_row})'
    sheet[f'{let}{max_row+1}'].style = 'Currency'
    
sheet["A1"] = 'Sales Report'
sheet["A2"] = month
sheet['A1'].font = Font('Times New Roman', bold=True, size=20)
sheet['A2'].font = Font('Times New Roman', italic=True,size=16)

output_path = os.path.join(app_path,f'report_{month}.xlsx')
print(output_path)
wb.save(output_path)